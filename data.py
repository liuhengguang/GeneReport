from utils import Params
from openpyxl import load_workbook
import jieba
import numpy as np


def xlsx_read(file_dir=Params.example_input, sheet_name="男（799）", min_col=0, max_col=14, min_row=2, max_row=75):
    wb = load_workbook(file_dir)
    sheet = wb[sheet_name]
    # print("===================表格基本信息：=======================")
    # print("sheet names:", wb.sheetnames)
    # print(sheet_name+"的维度：", sheet.calculate_dimension())
    r = sheet.calculate_dimension()
    print(r)
    data = sheet[r.split(":")[0][0]+":"+r.split(":")[1][0]]
    if max_row >= int(r.split(":")[1][1:]):
        max_row = int(r.split(":")[1][1:])
    total = []
    for col in range(min_col, max_col):  # 列
        temp = []
        for row in range(min_row, max_row):  # 行
            # print(data[row][col].value)
            temp.append(data[col][row].value)
        # print(temp)
        total.append(temp)
    # print(total)
    return total


def xlsx_read_v2(file_dir=Params.advices_regularization, sheet_name="yan_zheng_jiedu", min_col=0, max_col=1, min_row=0, max_row=50):
    wb = load_workbook(file_dir)
    sheet = wb[sheet_name]
    r = sheet.calculate_dimension()
    if max_row >= int(r.split(":")[1][1:]):
        max_row = int(r.split(":")[1][1:])
    # 通过指定范围(列 → 列)
    total = []
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        if row:
            for cell in row:
                # print(cell.value)
                total.append(cell.value)
    return total


def xlsx_read_v3(file_dir=Params.example_input, sheet_name="man_799", min_col=0, max_col=2, min_row=0, max_row=100):
    wb = load_workbook(file_dir)
    sheet = wb[sheet_name]
    r = sheet.calculate_dimension()
    if max_row >= int(r.split(":")[1][1:]):
        max_row = int(r.split(":")[1][1:])
    # 通过指定范围(列 → 列)
    regularized_input = []
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        if row:
            temp = []
            for cell in row:
                # print(cell.value)
                temp.append(cell.value)
            regularized_input.append(temp)
    return regularized_input


def get_dz_advices(file_dir=Params.advices_regularization):  # 获取邓总提供的建议
    wb = load_workbook(Params.advices_regularization)
    sheet_names = wb.sheetnames[0:-1]
    advices = dict()
    seg = dict()
    for name in sheet_names:
        ad = xlsx_read_v2(file_dir=file_dir, sheet_name=name, min_col=0, max_col=1, min_row=0, max_row=50)
        if ad[-1]:
            ad.append(None)
        temp_ad = []
        temp_seg = []
        cnt = 0
        for line in ad:
            if line:
                cnt = cnt + 1
                temp = ""
                # print(line)
                for i in line:
                    if i != " ":
                        temp = temp + i
                temp_ad.append(temp.split(".")[1])

            if not line:
                if cnt != 0:
                    temp_seg.append(cnt)
                cnt = 0
        advices[name]=temp_ad
        seg[name] = temp_seg

    return advices, seg


def cal_similarity(sentence1="", sentence2=""):
    sentence1 = jieba.cut(sentence1, cut_all=False)
    sentence2 = jieba.cut(sentence2, cut_all=False)
    s1 = []
    s2 = []
    for each in sentence1:
        s1.append(each)
    for each in sentence2:
        s2.append(each)

    s1 = ' '.join(str(i) for i in s1)
    s2 = ' '.join(str(i) for i in s2)
    word_bag = set(s1.split(" ")+s2.split(" "))

    s1_v = []
    for i in word_bag:
        if i in s1:
            s1_v.append(1)
        else:
            s1_v.append(0)

    s2_v = []
    for i in word_bag:
        if i in s2:
            s2_v.append(1)
        else:
            s2_v.append(0)

    fenzi = 0
    fenmu1 = 0
    fenmu2 = 0
    for i in range(0, len(word_bag)):
        fenzi = fenzi + s1_v[i]*s2_v[i]
        fenmu1 = fenmu1 + s1_v[i]*s1_v[i]
        fenmu2 = fenmu2 + s2_v[i]*s2_v[i]
    cos_dis = fenzi/(np.sqrt(fenmu1) * np.sqrt(fenmu2))

    return cos_dis, [s1, s1_v], [s2, s2_v], word_bag


if __name__ == "__main__":
    """
    advices, seg = get_dz_advices(file_dir=Params.advices_regularization)

    sen_freq = []
    advices_list = []
    for key in advices.keys():
        for line in advices[key]:
            advices_list.append(line)
            # print(line)

    # 按照句子之间的相似度进行分类
    classified_advices = []
    sorted_advices = sorted(advices_list)
    while(sorted_advices):
        temp = []
        query = sorted_advices[0]
        temp.append(query)
        remaining_sorted_advices = sorted_advices[1:]
        sorted_advices.remove(query)
        for i in range(0, len(remaining_sorted_advices)):
            cos_dis, s1, s2, word_bag = cal_similarity(query, remaining_sorted_advices[i])
            # print(cos_dis)
            if cos_dis >= 0.5:
                # print(remaining_sorted_advices[i])
                temp.append(remaining_sorted_advices[i])
                sorted_advices.remove(remaining_sorted_advices[i])

        classified_advices.append(temp)

    for l in classified_advices:
        print("\n")
        for line in l:
            print(line)
    """

    regularized_input = xlsx_read_v3(file_dir=Params.example_input, sheet_name="man_799", min_col=0, max_col=2, min_row=0, max_row=100)