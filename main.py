from utils import Params
from openpyxl import load_workbook
from advice_db.advice_db import *


def xlsx_read(file_dir=Params.example_file, sheet_name="男（799）", min_col=0, max_col=14, min_row=2, max_row=75):
    wb = load_workbook(file_dir)
    sheet = wb[sheet_name]
    # print("===================表格基本信息：=======================")
    # print("sheet names:", wb.sheetnames)
    print(sheet_name+"的维度：", sheet.calculate_dimension())
    r = sheet.calculate_dimension()
    data = sheet[r.split(":")[0][0]+":"+r.split(":")[1][0]]
    if max_row >= int(r.split(":")[1][1:]):
        max_row = int(r.split(":")[1][1:])
    total = []
    for col in range(min_col, max_col):  # 列
        temp = []
        for row in range(min_row, max_row):  # 行
            # print(data[row][col].value)
            temp.append(data[col][row].value)
        # print(temp)
        total.append(temp)
    # print(total)
    return total


def get_input(sheet_name="女（799）"):
    input_cols = xlsx_read(sheet_name=sheet_name)
    # a exp line in total_rows: (1, '骨科疾病', '骨关节炎', 0.609499828501506, 8.868, 14.5497, 2, None)]
    input_rows = list(zip(input_cols[0],
                          input_cols[1],
                          input_cols[2],
                          input_cols[6],
                          input_cols[7],
                          input_cols[9],
                          input_cols[12],
                          input_cols[13]
                          ))
    return input_cols, input_rows


def selecting_results():
    """
    从所有检测项目中挑出我们关注的项
    :return:
    """
    input_cols, input_rows = get_input()

    jibing_high_p = [i for i in input_rows
                     if i[0] == 1 and i[4] >= 10]
    jibing_high_population = [i for i in input_rows
                              if i[0] == 1 and i[3] >= 2 and i[4] <= 10]
    jibing = jibing_high_p + jibing_high_population

    kuangwz = [i[2] for i in input_rows
               if i[1] == "矿物质" and "高" in i[-1]]

    weiss = [i[2] for i in input_rows
             if i[1] == "维生素" and i[2] != "叶酸" and "高" in i[-1]]

    yesuan = [i[2] for i in input_rows
              if i[2] == "叶酸" and "高" in i[-1]]

    shansdx = [i[2] for i in input_rows
               if i[1] == "膳食代谢" and "高" in i[-1]]
    for i in input_rows:
        if i[1] == "膳食代谢" and ("浅尝" in i[-1] or "小酌" in i[-1]):
            shansdx.append("限酒")
        elif i[1] == "膳食代谢" and ("戒酒" in i[-1]):
            shansdx.append("戒酒")

    personal_features = [i for i in input_rows
                         if i[0] == 3]

    return jibing, kuangwz, weiss, yesuan, shansdx, personal_features


def knowledge():
    """
    step3: load knowledge
    :return: original_advices; splited_advices; splited_advices_v2; splited_advices_descriptions
    """
    advices_cols = xlsx_read(Params.man_and_woman_799_advices, sheet_name="man_and_woman_799", max_col=3, min_row=1, max_row=1000)
    knowledge_rows = list(zip(advices_cols[1], advices_cols[2]))
    cnt = 0
    cnt2 = 0
    diseases = []
    original_advices = dict()  # 原始建议
    splited_advices = dict()  # 分词结果
    label = False
    label2 = False
    if knowledge_rows[-1] != ("end", "end"):
        knowledge_rows.append(("end", "end"))
    for row in knowledge_rows:
        d = row[0]
        if "#" in d:
            temp = []
            label = True
            cnt = cnt + 1
            d = d[1:]
            d = d.replace(" ", "")
            diseases.append(d)
            # print(d)
        if label and "分词结果" in row[0]:
            # print(diseases[cnt-1])
            label = False
            original_advices[diseases[cnt-1]] = temp
        if label and "(" in row[0]:
            temp.append(row[0])

        if "分词结果" in d:
            temp2 = []
            label2 = True
            cnt2 = cnt2 + 1
            continue
            # print(d)
        if label2 and ("#" in row[0] or "end" in row[0]):
            # print(diseases[cnt-2])
            label2 = False
            splited_advices[diseases[cnt-2]] = temp2
        if label2:
            temp2.append((row[0], row[1]))

    # 建议描述
    splited_advices_descriptions = []
    for key in splited_advices.keys():
        for row in splited_advices[key]:
            splited_advices_descriptions.append(row[1])
            # print(row[1])
    splited_advices_descriptions = set(splited_advices_descriptions)  # 去处重复
    splited_advices_v2 = dict()
    for row in splited_advices_descriptions:
        temp = []
        for key in splited_advices.keys():
            for row1 in splited_advices[key]:
                if row1[1] == row:
                    temp.append(row1[0])
        splited_advices_v2[row] = temp
    return original_advices, splited_advices, splited_advices_v2, splited_advices_descriptions


# ================= #
# output advices
# ================= #
original_advices, splited_advices, splited_advices_v2, _ = knowledge()

jibing, kuangwz, weiss, yesuan, shansdx, personal_features = selecting_results()
diet_advices = []

# 仅仅根据疾病得出的初始建议
output_advices = []
for row in jibing:
    # print(row[2], ":")
    for advice in splited_advices[row[2]]:
        output_advices.append(advice)
        # print("    ", advice[0])
# 存储最终建议
final_output_advices = []
# 先加入各个疾病的私有建议
for row in output_advices:
    if "s" in row[1]:
        final_output_advices.append(row)
        output_advices.remove(row)
# 去除非私有建议中的重复建议
idx = [i[1] for i in output_advices]
idx = list(set(idx))

# 解决非私有建议与膳食建议中的冲突：
# 蛋白质：多蛋白质，少蛋白质； 维生素：多维生素C； 酒精：戒酒、限酒； 铁：多铁； 膳食纤维：多膳食纤维； 钙：避免过度补钙, 补钙
for i in idx:
    if "多维生素C" in i:
        weiss.append("维生素C")
        idx.remove(i)
    if "铁" in i:
        kuangwz.append("铁")
        idx.remove(i)
    if "酒" in i:
        shansdx.append(i)
        idx.remove(i)
weiss = list(set(weiss))
kuangwz = list(set(kuangwz))
shansdx = list(set(shansdx))
for row in idx:
    advice = splited_advices_v2[row][0]
    final_output_advices.append(advice)

# 膳食建议
diet_advices = []
# 矿物质相关建议
if kuangwz:
    for i in kuangwz:
        temp = kwz_advice_db[i]
        diet_advices.append(temp)
        # print(temp)

# 维生素相关建议
if weiss:
    label_B = False
    label_other = False
    other = []
    for i in weiss:
        if "B" in i:
            label_B = True
        elif "B" not in i and weiss:
            other.append(i)
            label_other = True

    if label_B and label_other:
        cnt = 0
        for v in other:
            if cnt == 0:
                vitamin_advice_db["B族和其他"] = vitamin_advice_db["B族和其他"] + other[cnt]
            else:
                vitamin_advice_db["B族和其他"] = vitamin_advice_db["B族和其他"] + "、" + other[cnt]
            cnt = cnt + 1
        diet_advices.append(vitamin_advice_db["B族和其他"])
        # print(vitamin_advice_db["B族和其他"])
    elif label_other and not label_B:
        cnt = 0
        for v in other:
            if cnt == 0:
                vitamin_advice_db["仅其他"] = vitamin_advice_db["仅其他"] + other[cnt]
            else:
                vitamin_advice_db["仅其他"] = vitamin_advice_db["仅其他"] + "、" + other[cnt]
            cnt = cnt + 1
        diet_advices.append(vitamin_advice_db["仅其他"])
        # print(vitamin_advice_db["仅其他"])
    elif not label_other and label_B:
        diet_advices.append(vitamin_advice_db["仅B族"])
        # print(vitamin_advice_db["仅B族"])
    else:
        print("None")
        pass

# 叶酸相关建议
if yesuan:
    diet_advices.append(vitamin_advice_db[yesuan[0]])

# 膳食代谢
label_jiejiu = False
label_xianjiu = False
for i in shansdx:
    if "乳糖" in i:
        diet_advices.append(rut[i])
    # 脂肪酸和膳食纤维
    if "膳食纤维" in i:
        diet_advices.append(shansxw[i])
    if "戒酒" in i:
        label_jiejiu = True
    if "限酒" in i:
        label_xianjiu = True
    if "咖啡因" in i:
        diet_advices.append(kafy["咖啡因"])
if label_jiejiu and label_xianjiu:
    diet_advices.append(jiuj["酒精"]["戒酒"])
elif not label_jiejiu and label_xianjiu:
    diet_advices.append(jiuj["酒精"]["限酒"])
else:
    pass
# 运动建议
personal_sport_features = []
for i in personal_features:
    if i[1] == "运动能力":
        personal_sport_features.append(i)

# output advices
for i in final_output_advices:
    print(i)
print("\n\n")
for i in diet_advices:
    print(i)






