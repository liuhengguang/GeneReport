from utils import Params
from openpyxl import load_workbook
from advice_db.advice_db import *


def xlsx_read(file_dir=Params.example_file, sheet_name="男（799）"):
    wb = load_workbook(file_dir)
    sheet = wb[sheet_name]
    # print("===================表格基本信息：=======================")
    # print("sheet names:", wb.sheetnames)
    # print("男（799）的维度：", sheet.calculate_dimension())
    data = sheet['A:N']
    total = []
    for col in range(0, 14):  # 列
        temp = []
        for row in range(2, 75):  # 行
            # print(data[row][col].value)
            temp.append(data[col][row].value)
        # print(temp)
        total.append(temp)
    # print(total)
    return total


# ======================================== #
def disease_advices(total):
    # ================= #
    # 疾病建议
    # ================= #
    advices = "根据您的遗传基因信息显示，您在:\n"
    diseases = []
    cnt = 0
    for i in range(0, 73):
        if total[0][i] == 1:
            if total[12][i]:
                if total[12][i] == 3:  # 3：患病概率高, 2:中, 1:低
                    cnt = cnt + 1
                    # print(advices)
                    if cnt == 1:
                        diseases.append(total[2][i])
                        advices = advices + total[2][i]
                    elif cnt > 1:
                        diseases.append(total[2][i])
                        advices = advices + "、" + total[2][i]
    advices = advices + "的患病概率较大\n您平时的体检需要特别注意以下内容\n"
    tijian = "318 套餐\n798 套餐\n1198 套餐\n1899 套餐\n3298 套餐\n(根据相关系数确定定制套餐内容)\n\n"
    advices = advices + tijian
    return advices, diseases
    # print(advices)


def diet_advices(total):
    # ================= #
    # 膳食建议
    # ================= #
    needs = {
        "矿物质": {},
        "维生素": {},
        "乳糖不耐受": {},
             }
    advices = "同时根据您的膳食需求可以看出,您对"
    for i in range(0, 73):
        if total[1][i] == "矿物质":
            if total[13][i]:
                cnt = 0
                if "高" in total[13][i]:
                    cnt = cnt + 1
                    if cnt == 1:
                        advices = advices + total[2][i]
                        needs["矿物质"][str(cnt-1)] = total[2][i]
                    elif cnt > 1:
                        advices = advices + "、" + total[2][i]
                        needs["矿物质"][str(cnt-1)] = total[2][i]
                    # print(total[2][i], ":", "需求高")

    cnt = 0
    for i in range(0, 73):
        if total[1][i] == "维生素":
            if total[13][i]:
                if "高" in total[13][i]:
                    advices = advices + "、" + total[2][i]
                    needs["维生素"][str(cnt)] = total[2][i]
                    cnt = cnt + 1
                    # print(total[2][i], ":", "需求高")
    advices = advices + "的需求较高"

    for i in range(0, 73):
        if total[2][i] == "乳糖":
            if total[13][i] == "乳糖不耐受":
                advices = advices + "以及" + "乳糖不耐受"
                needs["乳糖不耐受"][str(0)] = total[2][i]
                # print(total[13][i], ":\n")
                # print(rut["乳糖"])
    advices = advices + ",所以:"
    print(advices)

    label_B = False
    label_other = False
    other = []
    for key in needs.keys():
        if key == "矿物质":
            for subkey in needs[key]:
                print(kwz_advice_db[needs[key][subkey]])

        if key == "维生素":
            for subkey in needs[key]:
                if "B" in needs[key][subkey]:
                    label_B = True
                elif "B" not in needs[key][subkey] and needs[key][subkey]:
                    # print(needs[key][subkey])
                    other.append(needs[key][subkey])
                    label_other = True
                # print(vitamin_advice_db[needs[key][subkey]])

    if key == "乳糖不耐受":
        for subkey in needs[key]:
            print(rut[needs[key][subkey]])

    if "叶酸" in other:
        print(vitamin_advice_db["叶酸"])
        other.remove("叶酸")

    if label_B and label_other:
        cnt = 0
        for v in other:
            if cnt == 0:
                vitamin_advice_db["B族和其他"] = vitamin_advice_db["B族和其他"] + other[cnt]
            else:
                vitamin_advice_db["B族和其他"] = vitamin_advice_db["B族和其他"] + "、" + other[cnt]
            cnt = cnt + 1
        print(vitamin_advice_db["B族和其他"])
    elif label_other and not label_B:
        cnt = 0
        for v in other:
            if cnt == 0:
                vitamin_advice_db["仅其他"] = vitamin_advice_db["仅其他"] + other[cnt]
            else:
                vitamin_advice_db["仅其他"] = vitamin_advice_db["仅其他"] + "、" + other[cnt]
            cnt = cnt + 1
        print(vitamin_advice_db["仅其他"])
    elif not label_other and label_B:
        print(vitamin_advice_db["仅B族"])
    else:
        print("None")


def diet_advices_from_choosed_diseases(diseases):
    print(disease_advice_db[diseases[0]])


# ======================================== #
# main
# ======================================== #
total = xlsx_read(sheet_name="女（799）")
total_rows = list(zip(total[0],
                      total[1],
                      total[2],
                      total[6],
                      total[7],
                      total[9],
                      total[12],
                      total[13],
                      ))

advicese, diseases = disease_advices(total)
print(advicese)
diet_advices(total)

diet_advices_from_choosed_diseases(diseases)
print("\n\n"+sport_advices["通用1"])
print(sport_advices["通用2"])
print("\n\n"+social_advices["通用1"])
print(social_advices["通用2"])
print("\n\n(另外,如果您对具体某项遗传信息比较感兴趣需要了解更多内容, 建议您可以通过目录查找,如“高血压”在第 39 页.)")


